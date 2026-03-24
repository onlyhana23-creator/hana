#!/usr/bin/env python3
"""인당 객단가 2026 목표 — 단일 시트·단일 표 엑셀 생성."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 월: 2024.1 ~ 2026.2
Y2024 = [
    273_627.4304,
    252_266.7390,
    274_343.8468,
    270_663.6571,
    267_180.4149,
    258_923.7474,
    270_325.0318,
    270_508.5029,
    275_562.0150,
    292_883.0081,
    302_986.5106,
    291_184.2320,
]
Y2025 = [
    283_287.0339,
    278_598.7365,
    292_449.8527,
    293_654.8750,
    283_931.0581,
    282_801.8779,
    291_569.1923,
    276_695.9634,
    312_818.8704,
    276_041.1593,
    303_893.7483,
    310_542.5141,
]
Y2026_JF = [298_157.4325, 282_539.0333]

MONTHS_KO = [
    "1월",
    "2월",
    "3월",
    "4월",
    "5월",
    "6월",
    "7월",
    "8월",
    "9월",
    "10월",
    "11월",
    "12월",
]


def thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def main() -> None:
    out = Path(__file__).resolve().parent / "인당_객단가_2026목표_시뮬레이션.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "통합"

    r_jan_26 = Y2026_JF[0] / Y2025[0]
    r_feb_26 = Y2026_JF[1] / Y2025[1]
    g_mult = (r_jan_26 + r_feb_26) / 2
    g_pct = (g_mult - 1) * 100

    avg_2025 = sum(Y2025) / 12.0
    seas_2025 = [v / avg_2025 for v in Y2025]

    r_jan_25 = Y2025[0] / Y2024[0]
    r_feb_25 = Y2025[1] / Y2024[1]
    g_bt = (r_jan_25 + r_feb_25) / 2
    pred_25_mar_dec = [Y2024[m] * g_bt for m in range(2, 12)]
    act_25_mar_dec = [Y2025[m] for m in range(2, 12)]
    errs = [p - a for p, a in zip(pred_25_mar_dec, act_25_mar_dec)]
    mae = sum(abs(e) for e in errs) / len(errs)
    mape = sum(abs(e) / a for e, a in zip(errs, act_25_mar_dec)) / len(errs) * 100

    ncols = 11
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, "인당 객단가 — 월별 통합표 (시뮬레이션)")
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    param_text = (
        f"Step1: r1=26.01/25.01={r_jan_26:.6f}, r2=26.02/25.02={r_feb_26:.6f} → g=(r1+r2)/2={g_mult:.6f} (약 {g_pct:.2f}% YoY) | "
        f"Step2: A=2025년월평균={avg_2025:.4f}, 계절성지수=25년해당월/A | "
        f"Step3: 26년3~12월 목표=25년동월×g (1·2월은 실적) | "
        f"검증: g'=(25.01/24.01+25.02/24.02)/2={g_bt:.6f}, 25년3~12월 백테스트 MAE={mae:.4f}, MAPE={mape:.2f}%"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(2, 1, param_text)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 48

    headers = [
        "월",
        "2024_실적",
        "2025_실적",
        "2025_계절성지수\n(해당월÷A)",
        "배수_g\n(26대25 1·2월YoY평균)",
        "2026_목표·실적",
        "2026_구분",
        "백테스트_g'\n(25대24 1·2월YoY평균)",
        "백테스트_예측\n(24동월×g', 3~12월)",
        "백테스트_오차\n(예측-25실적)",
        "백테스트_APE%",
    ]
    header_row = 4
    for j, h in enumerate(headers, start=1):
        c = ws.cell(header_row, j, h)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="E8EEF7")
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = thin_border()

    for m in range(12):
        r = header_row + 1 + m
        target = Y2026_JF[m] if m < 2 else Y2025[m] * g_mult
        kind = "실적" if m < 2 else "목표(25×g)"
        if m >= 2:
            pred_bt = Y2024[m] * g_bt
            act = Y2025[m]
            err = pred_bt - act
            ape = abs(err) / act * 100 if act else None
        else:
            pred_bt = None
            err = None
            ape = None

        row_vals = [
            MONTHS_KO[m],
            Y2024[m],
            Y2025[m],
            seas_2025[m],
            g_mult,
            target,
            kind,
            g_bt,
            pred_bt,
            err,
            ape,
        ]
        for j, val in enumerate(row_vals, start=1):
            cell = ws.cell(r, j, val)
            cell.border = thin_border()
            if val is None:
                cell.value = "—"
            elif j in (2, 3, 4, 5, 6, 8, 9, 10, 11):
                if j == 4:
                    cell.number_format = "0.000000"
                elif j in (5, 8):
                    cell.number_format = "0.000000"
                elif j == 11:
                    cell.number_format = "0.00"
                else:
                    cell.number_format = "0.0000"

    sum_row = header_row + 13
    ws.cell(sum_row, 1, "합계·요약")
    ws.cell(sum_row, 1).font = Font(bold=True)
    total_2026 = sum(Y2026_JF) + sum(Y2025[m] * g_mult for m in range(2, 12))
    ws.cell(sum_row, 6, total_2026)
    ws.cell(sum_row, 6).number_format = "0.0000"
    ws.cell(sum_row, 6).font = Font(bold=True)
    ws.cell(sum_row, 7, "2026년 1~12월 합(1·2 실적+3~12 목표)")
    ws.merge_cells(start_row=sum_row, start_column=7, end_row=sum_row, end_column=ncols)
    for j in range(1, ncols + 1):
        ws.cell(sum_row, j).border = thin_border()

    note_row = sum_row + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols)
    ws.cell(
        note_row,
        1,
        "※ 계절성지수×A×g = 25년해당월×g 이므로 Step3 최종값은 2025 동월에 g를 곱한 것과 동일. 백테스트는 25년 3~12월만(연초 2개월 YoY로 연간 레벨 추정 가정의 적합성 점검).",
    )
    ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height = 36

    widths = [8, 14, 14, 14, 12, 16, 14, 14, 18, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out)
    print(f"Wrote: {out}")


if __name__ == "__main__":
    main()
